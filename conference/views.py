import logging
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from .models import Conference, FeedbackSurveyResponse, ReflectionSurveyResponse
from collections import Counter
from conference.functions import generate_otp

from conference.mails import send_otp_email
from conference.models import (
    Conference,
    OTPRequest,
    UserDetails,
    ConferenceOrganisers,
    ConferenceDetails,
    ConferenceRegistration,
)
from conference.utils import (
    export_to_excel,
    export_emails_for_newsletters,
    str_to_bool,
    str_to_int,
)


# Create your views here.
def home(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    conferences = Conference.objects.filter(is_published=True)
    conferenceDetails = ConferenceDetails.objects.filter(conference_id__in=conferences)
    context = {"conferences": conferences, "conference_details": conferenceDetails}
    return render(request, "conference/index.html", context=context)


def ludlogin(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    if request.method.lower() == "post":
        username = request.POST.get("email")
        password = request.POST.get("password")
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            messages.error(request, "Invalid credentials")
            return redirect("ludlogin")
    else:
        return render(request, "userauth/login.html")


def ludlogout(request):
    if request.user.is_authenticated:
        messages.success(request, "You are now logged out")
        logout(request)
    return redirect("home")


def ludregister(request):
    if request.method.lower() == "post":
        email = request.POST.get("email")
        otp = generate_otp()
        ret = send_otp_email(email, otp)
        if ret:
            messages.success(request, "Your OTP has been sent to you")
            if OTPRequest.objects.filter(email=email).exists():
                new_otp = OTPRequest.objects.get(email=email)
                new_otp.otp = otp
                new_otp.save()
            else:
                newotp = OTPRequest.objects.create(email=email, otp=otp)
                newotp.save()
            return redirect("ludregister_step_2", email)
        else:
            messages.error(
                request,
                "unable to process your request now,\
                            please try after some time or contacting LUD Admin",
            )
            return redirect("ludregister")
    return render(request, "userauth/registration.html")


def ludregister_step_2(request, email):
    if OTPRequest.objects.filter(email=email).exists():
        otpRequest = OTPRequest.objects.get(email=email)
        if request.method == "POST":
            otp = request.POST.get("otp")
            if otp == otpRequest.otp:
                messages.success(
                    request,
                    "Your OTP has been been validated, please create an account",
                )
                return redirect("ludregister_step_3", email)
            else:
                messages.error(request, "Invalid OTP")
                return redirect("ludregister_step_3", email)
        return render(
            request,
            "userauth/registration_step_2.html",
            context={"email": email, "otpRequest": otpRequest},
        )
    else:
        messages.error(request, "OTP Request does not exist, please try again")
        return redirect("ludregister")


def ludregister_step_3(request, email):
    if OTPRequest.objects.filter(email=email).exists():
        if request.method == "POST":
            first_name = request.POST.get("firstname")
            last_name = request.POST.get("lastname")
            gender = request.POST.get("gender")
            dob = request.POST.get("dob")
            designation = request.POST.get("designation")
            organization = request.POST.get("organization")
            mobile = request.POST.get("mobile")
            password = request.POST.get("password")
            newsletter = request.POST.get("opt_newsletter")
            if not newsletter:
                newsletter = False
            else:
                newsletter = True

            if User.objects.filter(email=email).exists():
                user = User.objects.get(email=email)
                user.set_password(password)
                user.save()
            else:
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    username=email,
                )
                user.save()

            if not UserDetails.objects.filter(user=user).exists():
                userdetails = UserDetails(
                    user=user,
                    gender=gender,
                    dob=dob,
                    designation=designation,
                    organization=organization,
                    mobile=mobile,
                    opt_newsletter=newsletter,
                )
                userdetails.save()

            try:
                conferences = ConferenceOrganisers.objects.filter(mails=user.email)
                if conferences.exists():
                    user.is_staff = True
                    user.save()
                    messages.success(
                        request,
                        "You are a conference organiser, your conferences are now linked",
                    )
            except:
                pass

            messages.success(request, "Your Account has been created")
            return redirect("ludlogin")

        return render(
            request, "userauth/registration_step_3.html", context={"email": email}
        )
    else:
        messages.error(request, "OTP Request does not exist, please try again")
        return redirect("ludregister")


def dashboard(request):
    if request.user.is_authenticated and request.user.is_superuser:
        active_conferences = Conference.objects.filter(is_published=True).count()
        all_conferences = Conference.objects.all().count()
        registrations = (
            Conference.objects.annotate(
                registration_count=Count("conferenceregistration")
            )
            .values("title", "registration_count", "is_published", "conference_id")
            .filter(is_published=True)
        )
        context = {
            "active_conferences": active_conferences,
            "all_conferences": all_conferences,
            "registrations": registrations,
        }
        return render(request, "siteadmin/dashboard.html", context=context)
    if request.user.is_authenticated and request.user.is_staff:
        registeredconference = ConferenceRegistration.objects.filter(
            user=request.user
        ).values("conference_id")
        conferences = Conference.objects.exclude(
            conference_id__in=registeredconference
        ).exclude(is_published=False)
        conferenceDetails = ConferenceDetails.objects.filter(
            conference_id__in=conferences
        )
        context = {"conferences": conferences, "conference_details": conferenceDetails}
        return render(request, "organiser/dashboard.html", context=context)
    if request.user.is_authenticated:
        registeredconference = ConferenceRegistration.objects.filter(
            user=request.user
        ).values("conference_id")
        conferences = Conference.objects.exclude(
            conference_id__in=registeredconference
        ).exclude(is_published=False)
        conferenceDetails = ConferenceDetails.objects.filter(
            conference_id__in=conferences
        )
        context = {
            "conferences": conferences,
            "registeredconference": registeredconference,
            "conference_details": conferenceDetails,
        }
        return render(request, "participant/dashboard.html", context=context)
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def adminconferencecreate(request):
    if request.user.is_authenticated and request.user.is_superuser:
        if request.method.lower() == "post":
            conference_title = request.POST.get("conferenceHeading")
            location = request.POST.get("location")
            venue = request.POST.get("venue")
            startDate = request.POST.get("fromdate")
            endDate = request.POST.get("enddate")
            organizer1 = request.POST.get("org1")
            organizer2 = request.POST.get("org2")
            organizer3 = request.POST.get("org3")
            org_mob_1 = request.POST.get("org_mob_1")
            org_mob_2 = request.POST.get("org_mob_2")
            org_mob_3 = request.POST.get("org_mob_3")
            gmaps_link = request.POST.get("gmaps")
            newconference = Conference(
                title=conference_title,
                location=location,
                venue=venue,
                start_date=startDate,
                end_date=endDate,
                organizer1=organizer1,
                organizer2=organizer2,
                organizer3=organizer3,
                mobile1=org_mob_1,
                mobile2=org_mob_2,
                mobile3=org_mob_3,
                google_maps_link=gmaps_link,
                created_by=request.user,
            )
            newconference.save()

            if not ConferenceOrganisers.objects.filter(
                mails=organizer1, conference=newconference
            ).exists():
                cof_org_1 = ConferenceOrganisers(
                    mails=organizer1, conference=newconference, created_by=request.user
                )
                cof_org_1.save()
            if not ConferenceOrganisers.objects.filter(
                mails=organizer2, conference=newconference
            ).exists():
                cof_org_2 = ConferenceOrganisers(
                    mails=organizer2, conference=newconference, created_by=request.user
                )
                cof_org_2.save()
            if not ConferenceOrganisers.objects.filter(
                mails=organizer3, conference=newconference
            ).exists():
                cof_org_3 = ConferenceOrganisers(
                    mails=organizer3, conference=newconference, created_by=request.user
                )
                cof_org_3.save()

            if User.objects.filter(username=organizer1).exists():
                user = User.objects.get(username=organizer1)
                user.is_staff = True
                user.save()

            if User.objects.filter(username=organizer2).exists():
                user = User.objects.get(username=organizer2)
                user.is_staff = True
                user.save()

            if User.objects.filter(username=organizer3).exists():
                user = User.objects.get(username=organizer3)
                user.is_staff = True
                user.save()

            messages.success(request, "Your conference has been created!")
            return redirect("admin_list_active_conference")
        return render(request, "siteadmin/newconference.html")
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def adminlistactiveconference(request):
    if request.user.is_authenticated and request.user.is_superuser:
        conferences = Conference.objects.filter(is_published=True).order_by(
            "-created_at"
        )
        return render(
            request,
            "siteadmin/activeconference.html",
            context={"conferences": conferences},
        )
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def adminlistcompletedconference(request):
    if request.user.is_authenticated and request.user.is_superuser:
        conferences = Conference.objects.filter(is_published=False).order_by(
            "-created_at"
        )
        return render(
            request,
            "siteadmin/listconferences.html",
            context={"conferences": conferences},
        )


def adminmanageconference(request, conference_id):
    if request.user.is_authenticated and request.user.is_superuser:
        conference = Conference.objects.get(pk=conference_id)
        try:
            conference_details = ConferenceDetails.objects.get(
                conference_id=conference_id
            )
        except:
            conference_details = None

        try:
            participant_count = ConferenceRegistration.objects.filter(
                conference_id=conference.conference_id
            ).count()
        except:
            participant_count = 0

        context = {
            "conference": conference,
            "conference_details": conference_details,
            "participant_count": participant_count,
        }
        return render(request, "siteadmin/conference_details.html", context=context)
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def adminconferencestatuschange(request, conference_id):
    if request.user.is_authenticated and request.user.is_superuser:
        conference = Conference.objects.get(pk=conference_id)
        conference.is_published = not conference.is_published
        conference.save()
        messages.success(request, "Your conference status has been updated")
        return redirect("admin_manage_conference", conference_id)
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def adminconferenceupdate(request, conference_id):
    if request.user.is_authenticated and request.user.is_superuser:
        conference = Conference.objects.get(pk=conference_id)

        if request.method.lower() == "post":
            conference_title = request.POST.get("conferenceHeading")
            location = request.POST.get("location")
            venue = request.POST.get("venue")
            startDate = request.POST.get("fromdate")
            endDate = request.POST.get("enddate")
            organizer1 = request.POST.get("org1")
            organizer2 = request.POST.get("org2")
            organizer3 = request.POST.get("org3")
            org_mob_1 = request.POST.get("org_mob_1")
            org_mob_2 = request.POST.get("org_mob_2")
            org_mob_3 = request.POST.get("org_mob_3")
            gmaps_link = request.POST.get("gmaps")
            banner = request.FILES.get("confbanner")
            brochure = request.FILES.get("confbrochure")
            theme = request.POST.get("conftheme")
            description = request.POST.get("confdesc")
            feedback = request.POST.get("conffeedback")
            social_facebook = request.POST.get("social_facebook")
            social_insta = request.POST.get("social_insta")
            social_linkedin = request.POST.get("social_linkedin")
            social_twitter = request.POST.get("social_twitter")
            social_youtube = request.POST.get("social_youtube")

            conference.title = conference_title
            conference.location = location
            conference.venue = venue
            conference.start_date = startDate
            conference.end_date = endDate
            conference.organizer1 = organizer1
            conference.organizer2 = organizer2
            conference.organizer3 = organizer3
            conference.mobile1 = org_mob_1
            conference.mobile2 = org_mob_2
            conference.mobile3 = org_mob_3
            conference.google_maps_link = gmaps_link
            conference.save()

            if not ConferenceOrganisers.objects.filter(
                mails=organizer1, conference=conference
            ).exists():
                cof_org_1 = ConferenceOrganisers(
                    mails=organizer1, conference=conference, created_by=request.user
                )
                cof_org_1.save()
            if not ConferenceOrganisers.objects.filter(
                mails=organizer2, conference=conference
            ).exists():
                cof_org_2 = ConferenceOrganisers(
                    mails=organizer2, conference=conference, created_by=request.user
                )
                cof_org_2.save()
            if not ConferenceOrganisers.objects.filter(
                mails=organizer3, conference=conference
            ).exists():
                cof_org_3 = ConferenceOrganisers(
                    mails=organizer3, conference=conference, created_by=request.user
                )
                cof_org_3.save()

            if User.objects.filter(username=organizer1).exists():
                user = User.objects.get(username=organizer1)
                user.is_staff = True
                user.save()

            if User.objects.filter(username=organizer2).exists():
                user = User.objects.get(username=organizer2)
                user.is_staff = True
                user.save()

            if User.objects.filter(username=organizer3).exists():
                user = User.objects.get(username=organizer3)
                user.is_staff = True
                user.save()

            if ConferenceDetails.objects.filter(
                conference_id=conference.conference_id
            ).exists():
                conferenceDetails = ConferenceDetails.objects.get(
                    conference_id=conference.conference_id
                )
                if banner:
                    conferenceDetails.conference_banner = banner
                if brochure:
                    conferenceDetails.conference_brochure = brochure
                if theme:
                    conferenceDetails.conference_theme = theme
                if description:
                    conferenceDetails.conference_description = description
                if feedback:
                    conferenceDetails.conference_feedback_link = feedback
                if social_facebook != "None":
                    conferenceDetails.social_facebook = social_facebook
                if social_insta != "None":
                    conferenceDetails.social_insta = social_insta
                if social_linkedin != "None":
                    conferenceDetails.social_linkedin = social_linkedin
                if social_twitter != "None":
                    conferenceDetails.social_twitter = social_twitter
                if social_youtube != "None":
                    conferenceDetails.social_youtube = social_youtube
                conferenceDetails.save()
            else:
                conferenceDetails = ConferenceDetails(
                    conference=conference,
                    conference_banner=banner,
                    conference_brochure=brochure,
                    conference_theme=theme,
                    conference_description=description,
                    conference_feedback_link=feedback,
                    social_facebook=social_facebook,
                    social_insta=social_insta,
                    social_linkedin=social_linkedin,
                    social_twitter=social_twitter,
                    social_youtube=social_youtube,
                )
                conferenceDetails.save()

            messages.success(request, "Your conference has been updated")
            return redirect("admin_conference_update", conference_id)

        try:
            conference_details = ConferenceDetails.objects.get(conference=conference)
        except ConferenceDetails.DoesNotExist:
            conference_details = None

        try:
            participant_count = ConferenceRegistration.objects.filter(
                conference_id=conference.conference_id
            ).count()
        except Exception:
            participant_count = 0

        context = {
            "conference": conference,
            "conference_details": conference_details,
            "participant_count": participant_count,
        }

        return render(request, "siteadmin/updateconference.html", context=context)
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def participatedconference(request):
    if request.user.is_authenticated:
        registeredconference = ConferenceRegistration.objects.filter(
            user=request.user
        ).values("conference_id")
        conferences = Conference.objects.filter(
            conference_id__in=registeredconference
        ).exclude(is_published=True)
        return render(
            request,
            "participant/listconferences.html",
            context={"conferences": conferences},
        )
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def registeredconference(request):
    if request.user.is_authenticated:
        registeredconference = ConferenceRegistration.objects.filter(
            user=request.user
        ).values("conference_id")
        conferences = Conference.objects.filter(
            conference_id__in=registeredconference
        ).exclude(is_published=False)
        return render(
            request,
            "participant/regconferences.html",
            context={"conferences": conferences},
        )
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def deregisteredconference(request, conference_id):
    if request.user.is_authenticated:
        registeredconference = ConferenceRegistration.objects.get(
            user=request.user, conference_id=conference_id
        )
        registeredconference.delete()
        messages.success(request, "Your conference registration has been deleted")
        return redirect("dashboard")
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


@login_required(login_url="ludlogin", redirect_field_name="next")
def participateconference(request, conference_id):
    if request.user.is_authenticated:
        conference = Conference.objects.get(pk=conference_id)
        if request.method.lower() == "post":
            interest = request.POST.get("participation")
            if ConferenceRegistration.objects.filter(
                conference_id=conference_id, user=request.user
            ).exists():
                conference_reg = ConferenceRegistration.objects.get(
                    user=request.user, conference_id=conference_id
                )
                conference_reg.interest = interest
                conference_reg.save()
            else:
                conferenceReg = ConferenceRegistration(
                    interest=interest,
                    conference_id=conference.conference_id,
                    user=request.user,
                )
                conferenceReg.save()
            messages.success(request, "You have registered for the conference")
            return redirect("registered_conference")
        try:
            conferenceDetails = ConferenceDetails.objects.get(
                conference_id=conference.conference_id
            )
        except:
            conferenceDetails = None
        context = {"conference": conference, "conferenceDetails": conferenceDetails}
        return render(
            request, "conference/conference_participation.html", context=context
        )
    else:
        messages.error(request, "You are not logged in")
        return redirect("ludlogin")


def conference_details(request, conference_id):
    conference = Conference.objects.get(pk=conference_id)
    try:
        conference_details = ConferenceDetails.objects.get(conference_id=conference_id)
    except:
        conference_details = None

    context = {"conference": conference, "conference_details": conference_details}
    return render(request, "conference/conference_details.html", context=context)


def conferencepass(request, conference_id):
    if request.user.is_authenticated:
        conference = Conference.objects.get(pk=conference_id)
        conferenceRegistration = ConferenceRegistration.objects.get(
            user=request.user, conference_id=conference.conference_id
        )
        try:
            conferenceDetails = ConferenceDetails.objects.get(
                conference_id=conference.conference_id
            )
        except:
            conferenceDetails = None
        context = {
            "conference": conference,
            "conferenceRegistration": conferenceRegistration,
            "conferenceDetails": conferenceDetails,
        }
        return render(request, "conference/conference_pass.html", context=context)
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def stafforganisingconferenes(request):
    if request.user.is_authenticated and request.user.is_staff:
        organisingConference = ConferenceOrganisers.objects.filter(
            mails=request.user.email
        ).values("conference_id")
        conferences = Conference.objects.filter(
            is_published=True, conference_id__in=organisingConference
        ).order_by("-created_at")
        return render(
            request,
            "organiser/newconference.html",
            context={"conferences": conferences},
        )
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def stafforganisedconferene(request):
    if request.user.is_authenticated and request.user.is_staff:
        organisingConference = ConferenceOrganisers.objects.filter(
            mails=request.user.email
        ).values("conference_id")
        conferences = Conference.objects.filter(
            is_published=False, conference_id__in=organisingConference
        ).order_by("-created_at")
        return render(
            request,
            "organiser/listconferences.html",
            context={"conferences": conferences},
        )
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def staffupdateconference(request, conference_id):
    if request.user.is_authenticated and request.user.is_staff:
        conference = Conference.objects.get(pk=conference_id)
        if request.method.lower() == "post":
            banner = request.FILES.get("confbanner")
            brochure = request.FILES.get("confbrochure")
            theme = request.POST.get("conftheme")
            description = request.POST.get("confdesc")
            feedback = request.POST.get("conffeedback")
            social_facebook = request.POST.get("social_facebook")
            social_insta = request.POST.get("social_insta")
            social_linkedin = request.POST.get("social_linkedin")
            social_twitter = request.POST.get("social_twitter")
            social_youtube = request.POST.get("social_youtube")

            if ConferenceDetails.objects.filter(
                conference_id=conference.conference_id
            ).exists():
                conferenceDetails = ConferenceDetails.objects.get(
                    conference_id=conference.conference_id
                )
                if banner:
                    conferenceDetails.conference_banner = banner
                if brochure:
                    conferenceDetails.conference_brochure = brochure
                if theme:
                    conferenceDetails.conference_theme = theme
                if description:
                    conferenceDetails.conference_description = description
                if feedback:
                    conferenceDetails.conference_feedback_link = feedback
                if social_facebook != "None":
                    conferenceDetails.social_facebook = social_facebook
                if social_insta != "None":
                    conferenceDetails.social_insta = social_insta
                if social_linkedin != "None":
                    conferenceDetails.social_linkedin = social_linkedin
                if social_twitter != "None":
                    conferenceDetails.social_twitter = social_twitter
                if social_youtube != "None":
                    conferenceDetails.social_youtube = social_youtube
                conferenceDetails.save()
            else:
                conferenceDetails = ConferenceDetails(
                    conference=conference,
                    conference_banner=banner,
                    conference_brochure=brochure,
                    conference_theme=theme,
                    conference_description=description,
                    conference_feedback_link=feedback,
                    social_facebook=social_facebook,
                    social_insta=social_insta,
                    social_linkedin=social_linkedin,
                    social_twitter=social_twitter,
                    social_youtube=social_youtube,
                )
                conferenceDetails.save()

            messages.success(request, "Your conference details has been updated")
            return redirect("staff_update_conference", conference_id)
        try:
            conferenceDetails = ConferenceDetails.objects.get(conference=conference)
        except:
            conferenceDetails = None
        context = {"conference": conference, "conferenceDetails": conferenceDetails}
        return render(request, "organiser/manage_conference.html", context)
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def download_registration_details(request, conference_id):
    if (
        request.user.is_authenticated
        and request.user.is_staff
        or request.user.is_superuser
    ):
        queryset = ConferenceRegistration.objects.filter(
            conference_id=conference_id
        ).values(
            "registration_date",
            "interest",
            "user__first_name",
            "user__last_name",
            "user__email",
            "user__userdetails__mobile",
            "user__userdetails__gender",
            "user__userdetails__dob",
            "user__userdetails__designation",
            "user__userdetails__organization",
        )
        # print(queryset)
        response = export_to_excel(queryset)
        return response
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def download_emails_for_newsletter(request):
    if (
        request.user.is_authenticated
        and request.user.is_staff
        or request.user.is_superuser
    ):
        queryset = UserDetails.objects.filter(opt_newsletter=True).values(
            "user__first_name", "user__last_name", "user__email"
        )
        response = export_emails_for_newsletters(queryset)
        return response
    else:
        messages.error(request, "You are not logged in")
        return redirect("home")


def one_time_participation(request, conference_id):
    if Conference.objects.filter(pk=conference_id).exists():
        conference = Conference.objects.get(pk=conference_id)
        if request.method == "POST":
            email = request.POST["email"]
            first_name = request.POST["firstname"]
            last_name = request.POST["lastname"]
            gender = request.POST["gender"]
            dob = request.POST["dob"]
            designation = request.POST["designation"]
            orgnization = request.POST["organization"]
            mobile = request.POST["mobile"]
            newsletter = False

            user = request.user
            if not User.objects.filter(email=email).exists():
                user = User.objects.create(
                    email=email,
                    password=mobile,
                    first_name=first_name,
                    last_name=last_name,
                    username=email,
                )
                user.save()
            else:
                user = User.objects.get(email=email)

            if not UserDetails.objects.filter(user=user).exists():
                userdetails = UserDetails(
                    user=user,
                    gender=gender,
                    dob=dob,
                    designation=designation,
                    organization=orgnization,
                    mobile=mobile,
                    opt_newsletter=newsletter,
                )
                userdetails.save()

            if ConferenceRegistration.objects.filter(
                conference_id=conference_id, user=user.pk
            ).exists():
                conference_reg = ConferenceRegistration.objects.get(
                    user=user, conference_id=conference_id
                )
                conference_reg.interest = "Attend-Onetime"
                conference_reg.save()
            else:
                conferenceReg = ConferenceRegistration(
                    interest="Attend-Onetime",
                    conference_id=conference.conference_id,
                    user=user,
                )
                conferenceReg.save()

            messages.success(request, "You have registered for the conference")
            return redirect("conference_details", conference_id)
        else:
            try:
                conference_details = ConferenceDetails.objects.get(
                    conference=conference.pk
                )
            except:
                conference_details = None
            return render(
                request,
                "conference/one_time_reg.html",
                context={
                    "conference": conference,
                    "conference_details": conference_details,
                },
            )
    else:
        messages.error(request, "Conference does not exist")
        return redirect("home")


def error_view_404(request, exception):
    return render(request, "userauth/error_page.html")


def error_view_500(request):
    return render(request, "userauth/error_page.html")


def str_to_bool(value):
    if value is None:
        return False
    truthy = {"true", "1", "yes", "y", "on"}
    return str(value).strip().lower() in truthy


def str_to_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def conference_toolkit(request):
    return render(request, "forms/conference_toolkit.html")


def feedback_survey(request):
    conferences = Conference.objects.filter(is_published=True)
    selected_conference = None

    if request.method == "POST":
        selected_id = request.POST.get("conference_id")
        if selected_id:
            selected_conference = get_object_or_404(
                Conference, conference_id=selected_id
            )
        else:
            messages.error(request, "Please select a conference.")
            return render(
                request, "forms/feedback_form.html", {"conferences": conferences}
            )

        if "full_name" in request.POST and selected_conference:
            engaging_activities = request.POST.getlist("engaging_activities")
            engaging_activities_str = (
                ", ".join(engaging_activities) if engaging_activities else None
            )

            user = request.user if request.user.is_authenticated else "None"

            FeedbackSurveyResponse.objects.create(
                user=user,
                conference=selected_conference,
                full_name=request.POST.get("full_name"),
                email=request.POST.get("email"),
                phone=request.POST.get("whatsapp") or None,
                age=request.POST.get("age") or None,
                gender=request.POST.get("gender") or None,
                occupation=request.POST.get("profession"),
                occupation_other=request.POST.get("profession_other") or None,
                location_type=request.POST.get("residence_type") or None,
                first_time=str_to_bool(request.POST.get("first_time")),
                q9_1=str_to_int(request.POST.get("q9_1")) or 3,
                q9_2=str_to_int(request.POST.get("q9_2")) or 3,
                q9_3=str_to_int(request.POST.get("q9_3")) or 3,
                q9_4=str_to_int(request.POST.get("q9_4")) or 3,
                q9_5=str_to_int(request.POST.get("q9_5")) or 3,
                q10_1=str_to_int(request.POST.get("q10_1")) or 3,
                q10_2=str_to_int(request.POST.get("q10_2")) or 3,
                q10_3=str_to_int(request.POST.get("q10_3")) or 3,
                q10_4=str_to_int(request.POST.get("q10_4")) or 3,
                q10_5=str_to_int(request.POST.get("q10_5")) or 3,
                q11_1=str_to_int(request.POST.get("q11_1")) or 3,
                q11_2=str_to_int(request.POST.get("q11_2")) or 3,
                q11_3=str_to_int(request.POST.get("q11_3")) or 3,
                q11_4=str_to_int(request.POST.get("q11_4")) or 3,
                q12_1=str_to_int(request.POST.get("q12_1")) or 3,
                q12_2=str_to_int(request.POST.get("q12_2")) or 3,
                q12_3=str_to_int(request.POST.get("q12_3")) or 3,
                q12_4=str_to_int(request.POST.get("q12_4")) or 3,
                q12_5=str_to_int(request.POST.get("q12_5")) or 3,
                followup_study=(
                    str_to_bool(request.POST.get("followup_study"))
                    if request.POST.get("followup_study")
                    else False
                ),
                satisfaction=str_to_int(request.POST.get("satisfaction")) or 3,
                takeaways=request.POST.get("takeaways"),
                suggestions=request.POST.get("suggestions"),
                team_interest=(
                    str_to_bool(request.POST.get("team_interest"))
                    if request.POST.get("team_interest")
                    else False
                ),
                engaging_activity=engaging_activities_str,
            )

            messages.success(request, "Thank you for your feedback!")
            return redirect("dashboard")

        else:
            messages.error(request, "Please fill in all required fields.")

    return render(
        request,
        "forms/feedback_form.html",
        {"conferences": conferences, "selected_conference": selected_conference},
    )


def reflection_survey(request):
    conferences = Conference.objects.filter(is_published=True)

    selected_conference = None
    if request.method == "POST":
        selected_id = request.POST.get("conference_id")
        if selected_id:
            selected_conference = Conference.objects.get(conference_id=selected_id)

        if "full_name" in request.POST:
            ReflectionSurveyResponse.objects.create(
                user=request.user,
                conference=selected_conference,
                full_name=request.POST.get("full_name"),
                email=request.POST.get("email"),
                occupation=request.POST.get("occupation"),
                occupation_other=request.POST.get("occupation_other") or "",
                connect_new=request.POST.get("connect_new"),
                stayed_in_touch=request.POST.get("stayed_in_touch"),
                opportunities_found=request.POST.get("opportunities_found"),
                motivated_to_volunteer=request.POST.get("motivated_to_volunteer"),
                participated_due_to_conf=request.POST.get("participated_due_to_conf"),
                engaged_in_theme=request.POST.get("engaged_in_theme"),
                improved_knowledge=request.POST.get("improved_knowledge"),
                philosophy_applied=request.POST.get("philosophy_applied"),
                more_informed=request.POST.get("more_informed"),
                leadership_enhanced=request.POST.get("leadership_enhanced"),
                more_socially_engaged=request.POST.get("more_socially_engaged"),
                more_socially_sensitive=request.POST.get("more_socially_sensitive"),
                making_impact=request.POST.get("making_impact"),
                key_takeaway=request.POST.get("key_takeaway"),
                recommend=request.POST.get("recommend") == "True",
                stay_involved=request.POST.get("stay_involved") == "True",
                org_willing_to_partner=request.POST.get("org_willing_to_partner")
                == "True",
            )
            messages.success(request, "Thank you for your reflection!")
            return redirect("dashboard")

    return render(
        request,
        "forms/reflection_survey.html",
        {"conferences": conferences, "selected_conference": selected_conference},
    )


def feedback_dashboard(request):

    conferences = Conference.objects.filter(is_published=True).order_by("title")
    selected_conference_id = request.GET.get("conference_id")

    responses = FeedbackSurveyResponse.objects.all().order_by("-submitted_at")
    if selected_conference_id:
        responses = responses.filter(conference_id=selected_conference_id)
    total_responses = responses.count()

    # Map of model field names to chart display titles
    likert_chart_fields = {
        # Q9 – Volunteerism & Community Engagement
        "q9_1": "Interest in Community Volunteering",
        "q9_2": "Sense of Belonging to Community",
        "q9_3": "Inspired by Speakers/Stories",
        "q9_4": "Stalls/Interactions Offered Practical Insights",
        "q9_5": "Encouraged Reflection on Community Role",
        # Q10 – Social Entrepreneurship & Community Service
        "q10_1": "Confidence to Start an Initiative",
        "q10_2": "Showcased Practical Service Models",
        "q10_3": "Motivated to Volunteer/Serve",
        "q10_4": "Increased Awareness of Social Issues",
        "q10_5": "Connected with Organizations/Opportunities",
        # Q11 – Learning & Knowledge Sharing
        "q11_1": "Learned New Concepts/Approaches",
        "q11_2": "Better Understanding of Community Needs",
        "q11_3": "Applicable Learnings to Context",
        "q11_4": "Improved Problem-Solving/Decision Making",
        # Q12 – Networking & Collaborations
        "q12_1": "Met People with Similar Goals",
        "q12_2": "Gained Mentorship/Guidance",
        "q12_3": "Found Collaboration Opportunities",
        "q12_4": "Facilitated New Partnerships",
        "q12_5": "Strengthened Existing Relationships",
        # Final
        "followup_study": "Interest in Follow-Up Study (Yes/No)",
        "satisfaction": "Overall Satisfaction (1–5)",
    }

    charts_data = []

    # Location of participants chart
    location_counts = Counter(
        res.location_type for res in responses if res.location_type
    )
    if location_counts:
        charts_data.append(
            {
                "chart_id": "locationChart",
                "title": "Location of Participants",
                "labels": list(location_counts.keys()),
                "data": list(location_counts.values()),
            }
        )

    # Generate charts for Likert scale / boolean fields
    for field, title in likert_chart_fields.items():
        counts = Counter(
            getattr(res, field)
            for res in responses
            if hasattr(res, field) and getattr(res, field) is not None
        )
        if counts:
            keys = (
                sorted(counts.keys())
                if not isinstance(next(iter(counts.keys())), bool)
                else [False, True]
            )
            labels = (
                ["No", "Yes"]
                if keys and isinstance(keys[0], bool)
                else [str(k) for k in keys]
            )
            data = [counts.get(k, 0) for k in keys]
            charts_data.append(
                {
                    "chart_id": f"{field}Chart",
                    "title": title,
                    "labels": labels,
                    "data": data,
                }
            )

    # Engaging Activities chart: split comma-separated values and count each
    engaging_counts = Counter()
    for res in responses:
        if res.engaging_activity:
            activities = [
                act.strip() for act in res.engaging_activity.split(",") if act.strip()
            ]
            engaging_counts.update(activities)

    if engaging_counts:
        charts_data.append(
            {
                "chart_id": "engagingActivitiesChart",
                "title": "Most Engaging Activities",
                "labels": list(engaging_counts.keys()),
                "data": list(engaging_counts.values()),
            }
        )

    return render(
        request,
        "forms/feedback_dashboard.html",
        {
            "responses": responses,
            "total_responses": total_responses,
            "charts_data": charts_data,
            "conferences": conferences,
            "selected_conference_id": selected_conference_id or "",
        },
    )


def reflection_dashboard(request):

    conferences = Conference.objects.filter(is_published=True).order_by("title")
    selected_conference_id = request.GET.get("conference_id")

    responses = ReflectionSurveyResponse.objects.all().order_by("-submitted_at")
    if selected_conference_id:
        responses = responses.filter(conference_id=selected_conference_id)
    total = responses.count()

    chart_fields = {
        "connect_new": "Connected with New People",
        "stayed_in_touch": "Stayed in Touch Post Conference",
        "opportunities_found": "Found Collaboration Opportunities",
        "motivated_to_volunteer": "Motivated to Volunteer",
        "participated_due_to_conf": "Participated Because of Conference",
        "engaged_in_theme": "Engaged in Conference Themes",
        "improved_knowledge": "Improved Knowledge",
        "philosophy_applied": "Conference Philosophy Still Applies",
        "more_informed": "More Informed After Conference",
        "leadership_enhanced": "Enhanced Leadership Skills",
        "more_socially_engaged": "More Socially Engaged",
        "more_socially_sensitive": "More Socially Sensitive",
        "making_impact": "Believes Making an Impact",
        "recommend": "Would Recommend the Conference",
        "stay_involved": "Wants to Stay Involved",
        "org_willing_to_partner": "Willing to Partner with LUD",
    }

    charts_data = []
    for field, title in chart_fields.items():
        counts = Counter(
            getattr(res, field)
            for res in responses
            if hasattr(res, field) and getattr(res, field) is not None
        )
        if counts:
            labels = (
                ["Yes", "No"]
                if isinstance(list(counts.keys())[0], bool)
                else [str(k) for k in sorted(counts.keys())]
            )
            data = (
                [counts.get(True, 0), counts.get(False, 0)]
                if isinstance(list(counts.keys())[0], bool)
                else [counts[k] for k in sorted(counts.keys())]
            )
            charts_data.append(
                {
                    "chart_id": f"{field}Chart",
                    "title": title,
                    "labels": labels,
                    "data": data,
                }
            )

    return render(
        request,
        "forms/reflection_dashboard.html",
        {
            "responses": responses,
            "total_responses": total,
            "charts_data": charts_data,
            "conferences": conferences,
            "selected_conference_id": selected_conference_id or "",
        },
    )

def download_feedback_survey(request, conference_id):
    """Download feedback survey responses as Excel file"""
    if not request.user.is_superuser:
        messages.error(request, "You don't have permission to download this report.")
        return redirect('dashboard')
    
    conference = get_object_or_404(Conference, conference_id=conference_id)
    responses = FeedbackSurveyResponse.objects.filter(
        conference=conference
    ).order_by('-submitted_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Survey"
    
    # Define headers
    headers = [
        'Submission Date', 'Full Name', 'Email', 'Phone', 'Age', 'Gender',
        'Occupation', 'Occupation Other', 'Location Type', 'First Time',
        'Q9.1 (Volunteering Interest)', 'Q9.2 (Sense of Belonging)',
        'Q9.3 (Inspired by Speakers)', 'Q9.4 (Practical Insights)',
        'Q9.5 (Reflection on Role)', 'Q10.1 (Confidence to Start)',
        'Q10.2 (Service Models)', 'Q10.3 (Motivated to Serve)',
        'Q10.4 (Social Issues Awareness)', 'Q10.5 (Connected with Orgs)',
        'Q11.1 (Learned New Concepts)', 'Q11.2 (Understanding Needs)',
        'Q11.3 (Applicable Learnings)', 'Q11.4 (Problem Solving)',
        'Q12.1 (Met Similar Goals)', 'Q12.2 (Gained Mentorship)',
        'Q12.3 (Collaboration Opportunities)', 'Q12.4 (New Partnerships)',
        'Q12.5 (Strengthened Relationships)', 'Follow-up Study Interest',
        'Overall Satisfaction', 'Engaging Activities', 'Key Takeaways',
        'Suggestions', 'Team Interest'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row_num, response in enumerate(responses, 2):
        ws.cell(row=row_num, column=1, value=response.submitted_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=2, value=response.full_name)
        ws.cell(row=row_num, column=3, value=response.email)
        ws.cell(row=row_num, column=4, value=response.phone or '')
        ws.cell(row=row_num, column=5, value=response.age or '')
        ws.cell(row=row_num, column=6, value=response.gender or '')
        ws.cell(row=row_num, column=7, value=response.occupation or '')
        ws.cell(row=row_num, column=8, value=response.occupation_other or '')
        ws.cell(row=row_num, column=9, value=response.location_type or '')
        ws.cell(row=row_num, column=10, value='Yes' if response.first_time else 'No')
        ws.cell(row=row_num, column=11, value=response.q9_1)
        ws.cell(row=row_num, column=12, value=response.q9_2)
        ws.cell(row=row_num, column=13, value=response.q9_3)
        ws.cell(row=row_num, column=14, value=response.q9_4)
        ws.cell(row=row_num, column=15, value=response.q9_5)
        ws.cell(row=row_num, column=16, value=response.q10_1)
        ws.cell(row=row_num, column=17, value=response.q10_2)
        ws.cell(row=row_num, column=18, value=response.q10_3)
        ws.cell(row=row_num, column=19, value=response.q10_4)
        ws.cell(row=row_num, column=20, value=response.q10_5)
        ws.cell(row=row_num, column=21, value=response.q11_1)
        ws.cell(row=row_num, column=22, value=response.q11_2)
        ws.cell(row=row_num, column=23, value=response.q11_3)
        ws.cell(row=row_num, column=24, value=response.q11_4)
        ws.cell(row=row_num, column=25, value=response.q12_1)
        ws.cell(row=row_num, column=26, value=response.q12_2)
        ws.cell(row=row_num, column=27, value=response.q12_3)
        ws.cell(row=row_num, column=28, value=response.q12_4)
        ws.cell(row=row_num, column=29, value=response.q12_5)
        ws.cell(row=row_num, column=30, value='Yes' if response.followup_study else 'No')
        ws.cell(row=row_num, column=31, value=response.satisfaction)
        ws.cell(row=row_num, column=32, value=response.engaging_activity or '')
        ws.cell(row=row_num, column=33, value=response.takeaways or '')
        ws.cell(row=row_num, column=34, value=response.suggestions or '')
        ws.cell(row=row_num, column=35, value='Yes' if response.team_interest else 'No')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Feedback_Survey_{conference.title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def download_reflection_survey(request, conference_id):
    """Download reflection survey responses as Excel file"""
    if not request.user.is_superuser:
        messages.error(request, "You don't have permission to download this report.")
        return redirect('dashboard')
    
    conference = get_object_or_404(Conference, conference_id=conference_id)
    responses = ReflectionSurveyResponse.objects.filter(
        conference=conference
    ).order_by('-submitted_at')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reflection Survey"
    
    # Define headers
    headers = [
        'Submission Date', 'Full Name', 'Email', 'Occupation', 'Occupation Other',
        'Connected with New People', 'Stayed in Touch', 'Found Opportunities',
        'Motivated to Volunteer', 'Participated Due to Conference',
        'Engaged in Theme', 'Improved Knowledge', 'Philosophy Applied',
        'More Informed', 'Leadership Enhanced', 'More Socially Engaged',
        'More Socially Sensitive', 'Making Impact', 'Key Takeaway',
        'Would Recommend', 'Stay Involved', 'Willing to Partner'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row_num, response in enumerate(responses, 2):
        ws.cell(row=row_num, column=1, value=response.submitted_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_num, column=2, value=response.full_name)
        ws.cell(row=row_num, column=3, value=response.email)
        ws.cell(row=row_num, column=4, value=response.occupation or '')
        ws.cell(row=row_num, column=5, value=response.occupation_other or '')
        ws.cell(row=row_num, column=6, value=response.connect_new or '')
        ws.cell(row=row_num, column=7, value=response.stayed_in_touch or '')
        ws.cell(row=row_num, column=8, value=response.opportunities_found or '')
        ws.cell(row=row_num, column=9, value=response.motivated_to_volunteer or '')
        ws.cell(row=row_num, column=10, value=response.participated_due_to_conf or '')
        ws.cell(row=row_num, column=11, value=response.engaged_in_theme or '')
        ws.cell(row=row_num, column=12, value=response.improved_knowledge or '')
        ws.cell(row=row_num, column=13, value=response.philosophy_applied or '')
        ws.cell(row=row_num, column=14, value=response.more_informed or '')
        ws.cell(row=row_num, column=15, value=response.leadership_enhanced or '')
        ws.cell(row=row_num, column=16, value=response.more_socially_engaged or '')
        ws.cell(row=row_num, column=17, value=response.more_socially_sensitive or '')
        ws.cell(row=row_num, column=18, value=response.making_impact or '')
        ws.cell(row=row_num, column=19, value=response.key_takeaway or '')
        ws.cell(row=row_num, column=20, value='Yes' if response.recommend else 'No')
        ws.cell(row=row_num, column=21, value='Yes' if response.stay_involved else 'No')
        ws.cell(row=row_num, column=22, value='Yes' if response.org_willing_to_partner else 'No')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Reflection_Survey_{conference.title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response